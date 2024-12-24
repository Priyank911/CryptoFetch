import requests
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import os

def get_crypto_data(num_coins=50):
    url = f"https://api.coingecko.com/api/v3/coins/markets?vs_currency=usd&order=market_cap_desc&per_page={num_coins}&page=1&sparkline=false"
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error: {e}")
        return None

def analyze(data):
    if data is None:
        print("No data to analyze.")
        return None

    df = pd.DataFrame(data)
    top5 = df.nlargest(5, 'market_cap')
    avg_price = df['current_price'].mean()
    highest = df.nlargest(1, 'price_change_percentage_24h')
    lowest = df.nsmallest(1, 'price_change_percentage_24h')
    
    return {
        'top5': top5[['name', 'symbol', 'market_cap']].to_dict('records'),
        'avg_price': avg_price,
        'highest': highest[['name','symbol', 'price_change_percentage_24h']].to_dict('records'),
        'lowest': lowest[['name', 'symbol','price_change_percentage_24h']].to_dict('records')
    }

def update_excel(data, analysis, filename='crypto_live_data.xlsx'):
    if not data:
        print("No data for Excel.")
        return

    df = pd.DataFrame(data)
    df_excel = df[['market_cap_rank', 'name', 'symbol', 'current_price', 'market_cap', 'price_change_percentage_24h', 'total_volume']]
    
    for col in df_excel.columns:
        df_excel.loc[:, col] = df_excel[col].apply(lambda x: str(x) if isinstance(x, (dict, list)) else x)
    df_excel = df_excel.fillna('')
    
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(list(df_excel.columns))
    else:
        wb = load_workbook(filename)
        ws = wb.active
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

    for row in df_excel.values.tolist():
        ws.append(row)

    try:
        wb.save(filename)
        print("Excel updated.")
    except PermissionError:
        print(f"Error: Could not save {filename}. File might be open.")

def main():
    while True:
        print("\nUpdating data...")
        crypto_data = get_crypto_data()
        if crypto_data:
            analysis = analyze(crypto_data)
            if analysis:
                update_excel(crypto_data,analysis)

                print("\nAnalysis Report:")
                print("  Top 5 by Market Cap:")
                for item in analysis['top5']:
                    print(f"    - {item['name']} ({item['symbol']}): ${item['market_cap']:,}")
                print(f"  Average Price: ${analysis['avg_price']:.2f}")
                print("  Highest 24h Change:")
                for item in analysis['highest']:
                    print(f"    - {item['name']} ({item['symbol']}): {item['price_change_percentage_24h']:.2f}%")
                print("  Lowest 24h Change:")
                for item in analysis['lowest']:
                    print(f"    - {item['name']} ({item['symbol']}): {item['price_change_percentage_24h']:.2f}%")
            else:
                print("Analysis failed.")
        else:
            print("Data fetch failed.")
            
        print("Waiting 5 minutes...")
        time.sleep(300)

if __name__ == "__main__":
    main()